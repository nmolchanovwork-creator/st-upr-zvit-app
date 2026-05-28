import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Настройка страницы
st.set_page_config(page_title="Генератор Упр Звітів", layout="wide")
st.title("📊 Сквозна аналітика: Управлінський звіт з 1С")
st.write("Завантажте журнал проводок, і система автоматично збере шахматку по всім номенклатурним групам.")

# Виджет загрузки файла
uploaded_file = st.file_uploader("📥 Завантажте файл журналу проводок (.xlsx, .xls або .csv)", type=['xlsx', 'xls', 'csv'])

if uploaded_file is not None:
    filename = uploaded_file.name
    st.info(f"⏳ Файл '{filename}' завантажено. Йде аналіз даних...")

    try:
        if filename.endswith('.csv') or filename.endswith('.txt'):
            df = pd.read_csv(uploaded_file, sep='\t', dtype=str)
        else:
            df = pd.read_excel(uploaded_file, dtype=str)
    except Exception as e:
        st.error(f"Помилка при читанні файлу: {e}")
        df = pd.DataFrame()

    if not df.empty:
        df.columns = [col.strip() for col in df.columns]

        sum_col = 'Сума' if 'Сума' in df.columns else [c for c in df.columns if 'сума' in c.lower() and '(' not in c.lower()][0]
        df['NumericSum'] = pd.to_numeric(df[sum_col], errors='coerce').fillna(0)

        unique_groups = set()
        for _, row in df.iterrows():
            kt_acc = str(row.get('Рахунок Кт', '')).strip()
            dt_acc = str(row.get('Рахунок Дт', '')).strip()
            sub2_kt = str(row.get('Субконто2 Кт', '')).strip()
            sub1_dt = str(row.get('Субконто1 Дт', '')).strip()

            invalid_vals = ['nan', '<...>', 'none', '']

            if kt_acc.startswith('703') and sub2_kt.lower() not in invalid_vals:
                unique_groups.add(sub2_kt)
            if dt_acc.startswith('903') and sub1_dt.lower() not in invalid_vals:
                unique_groups.add(sub1_dt)

        groups_list = sorted(list(unique_groups))

        if not groups_list:
            st.warning("⚠️ В файлі не знайдено жодної номенклатурної групи по рахункам 703 або 903.")
        else:
            st.success(f"⚙️ Знайдено {len(groups_list)} проектів. Формуємо звіт...")

            # Створення Excel-файлу в пам'яті
            wb = Workbook()
            ws = wb.active
            ws.title = "Сводный Упр Звіт"
            ws.views.sheetView[0].showGridLines = False

            f_bold = Font(name='Arial', bold=True, size=10)
            f_header = Font(name='Arial', bold=True, size=11, color='FFFFFF')
            f_group_title = Font(name='Arial', bold=True, size=12, color='FFFFFF')

            fill_hdr = PatternFill('solid', fgColor='2E4057')
            fill_group = PatternFill('solid', fgColor='1F4E79')
            fill_separator = PatternFill('solid', fgColor='E0E0E0')

            border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                            top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))

            headers = ["Номенклатурная группа", "Документы", "Дата", "Контрагент", "Заход денег 311/361", "Расходы 231/903", "Продажа 361/703"]
            for col_idx, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col_idx, value=h)
                c.font = f_header; c.fill = fill_hdr; c.alignment = Alignment(horizontal='center', vertical='center')

            ws.row_dimensions[1].height = 25
            curr_row = 3

            # --- ПРОГРЕСС БАР ---
            progress_bar = st.progress(0)
            
            for i, target_group in enumerate(groups_list):
                
                # Обновление прогресс-бара
                progress_bar.progress((i + 1) / len(groups_list))

                linked_clients = set()
                for _, row in df.iterrows():
                    dt_acc = str(row.get('Рахунок Дт', '')).strip()
                    kt_acc = str(row.get('Рахунок Кт', '')).strip()
                    sub1_dt, sub2_dt = str(row.get('Субконто1 Дт', '')), str(row.get('Субконто2 Дт', ''))
                    sub1_kt, sub2_kt = str(row.get('Субконто1 Кт', '')), str(row.get('Субконто2 Кт', ''))

                    if kt_acc.startswith('703') and target_group in sub2_kt:
                        client = sub1_dt.strip()
                        if client and client != '<...>': linked_clients.add(client)
                    elif dt_acc.startswith('903') and target_group in sub1_dt:
                        client = sub1_kt.strip()
                        if client and client != '<...>': linked_clients.add(client)

                if not linked_clients:
                    linked_clients.add(target_group)

                bank_data, expense_data, sale_data = [], [], []

                for _, row in df.iterrows():
                    dt_acc = str(row.get('Рахунок Дт', '')).strip()
                    kt_acc = str(row.get('Рахунок Кт', '')).strip()
                    v_sum = float(row['NumericSum'])
                    if v_sum == 0: continue

                    date_v = str(row.get('Період', '')).split(' ')[0]
                    content = str(row.get('Зміст', ''))
                    registrar = str(row.get('Реєстратор', '')).lower()

                    sub_dt = [str(row.get('Субконто1 Дт', '')), str(row.get('Субконто2 Дт', '')), str(row.get('Субконто3 Дт', ''))]
                    sub_kt = [str(row.get('Субконто1 Кт', '')), str(row.get('Субконто2 Кт', '')), str(row.get('Субконто3 Кт', ''))]

                    if 'закриття місяця' in registrar:
                        continue

                    if dt_acc.startswith('31') and kt_acc.startswith('361'):
                        client = sub_kt[0].strip()
                        if any(c in client for c in linked_clients) or target_group in content:
                            bank_data.append(["Банк (Оплата)", date_v, client, v_sum, "", ""])

                    elif dt_acc.startswith('361') and kt_acc.startswith('31'):
                        client = sub_dt[0].strip()
                        if any(c in client for c in linked_clients) or target_group in content:
                            bank_data.append(["Банк (Возврат)", date_v, client, -v_sum, "", ""])

                    elif dt_acc.startswith('361') and kt_acc.startswith('703'):
                        client = sub_dt[0].strip()
                        if any(c in client for c in linked_clients) or target_group in sub_kt[1]:
                            sale_data.append(["Реалізація", date_v, client, "", "", v_sum])

                    elif (dt_acc.startswith('231') or dt_acc.startswith('903')) and (target_group in sub_dt[0] or target_group in sub_dt[1]):
                        supplier = "Первинні витрати (Списання)"
                        for s in sub_kt:
                            if s and s != '<...>' and 'поточний' not in s.lower() and s != 'nan':
                                supplier = s; break
                        expense_data.append(["Покупка (Витрати)", date_v, supplier, "", v_sum, ""])

                sum_bank = sum([r[3] for r in bank_data])
                sum_exp = sum([r[4] for r in expense_data])
                sum_sale = sum([r[5] for r in sale_data])
                margin = sum_sale - sum_exp
                margin_pct = (margin / sum_sale) if sum_sale > 0 else 0

                ws.cell(row=curr_row, column=1, value=target_group).font = f_group_title
                ws.cell(row=curr_row, column=1).fill = fill_group
                for c in range(2, 8): ws.cell(row=curr_row, column=c).fill = fill_group
                curr_row += 1

                for r in bank_data:
                    for col_idx, val in enumerate(r, 2):
                        c = ws.cell(row=curr_row, column=col_idx, value=val)
                        c.border = border
                        if col_idx == 5 and isinstance(val, (int, float)): c.number_format = '#,##0.00'
                    curr_row += 1

                ws.cell(row=curr_row, column=2, value="итого заход денег").font = f_bold
                ws.cell(row=curr_row, column=5, value=sum_bank).font = f_bold
                ws.cell(row=curr_row, column=5).number_format = '#,##0.00'
                curr_row += 2

                for r in expense_data:
                    for col_idx, val in enumerate(r, 2):
                        c = ws.cell(row=curr_row, column=col_idx, value=val)
                        c.border = border
                        if col_idx == 6 and isinstance(val, (int, float)): c.number_format = '#,##0.00'
                    curr_row += 1

                ws.cell(row=curr_row, column=2, value="итого покупка").font = f_bold
                ws.cell(row=curr_row, column=6, value=sum_exp).font = f_bold
                ws.cell(row=curr_row, column=6).number_format = '#,##0.00'
                curr_row += 2

                for r in sale_data:
                    for col_idx, val in enumerate(r, 2):
                        c = ws.cell(row=curr_row, column=col_idx, value=val)
                        c.border = border
                        if col_idx == 7 and isinstance(val, (int, float)): c.number_format = '#,##0.00'
                    curr_row += 1

                ws.cell(row=curr_row, column=7, value=sum_sale).font = f_bold
                ws.cell(row=curr_row, column=7).number_format = '#,##0.00'
                curr_row += 2

                ws.cell(row=curr_row, column=2, value="маржа").font = f_bold
                ws.cell(row=curr_row, column=7, value=margin).font = f_bold
                ws.cell(row=curr_row, column=7).number_format = '#,##0.00'

                curr_row += 1
                ws.cell(row=curr_row, column=2, value="маржа %").font = f_bold
                ws.cell(row=curr_row, column=7, value=margin_pct).font = f_bold
                ws.cell(row=curr_row, column=7).number_format = '0.0%'

                curr_row += 2
                for c in range(1, 8): ws.cell(row=curr_row, column=c).fill = fill_separator
                curr_row += 4

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col[:100])
                ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 3, 16)
            ws.column_dimensions['A'].width = 35

            # Сохранение файла в оперативную память (для веба)
            output = io.BytesIO()
            wb.save(output)
            processed_data = output.getvalue()

            st.balloons() # Праздничные шарики по завершению!
            
            # Кнопка для скачивания
            st.download_button(
                label="📥 Завантажити готовий Управлінський Звіт",
                data=processed_data,
                file_name="Упр_Звіт_Всі_Проєкти.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
